#!/usr/bin/env python3
# -*- coding:utf-8 -*-

# @Time    : 2025/11/10 16:30
# @Author  : kaichuan
# @Email   : zhaoweichen.zwc@antgroup.com
# @FileName: excel_tool.py

import os
import json
import asyncio
from typing import Optional, List, Dict, Any, Union
from enum import Enum
from dataclasses import dataclass
from pydantic import Field

from agentuniverse.agent.action.tool.tool import Tool
from agentuniverse.base.annotation.retry import retry
from agentuniverse.base.util.logging.logging_util import LOGGER


class ExcelMode(Enum):
    """Excel operation modes"""
    READ = "read"
    WRITE = "write"
    APPEND = "append"
    INFO = "info"


@dataclass
class ExcelInfo:
    """Excel file information"""
    file_path: str
    file_size: int
    sheets: List[str]
    total_sheets: int
    primary_sheet: str
    row_count: Dict[str, int]
    col_count: Dict[str, int]


class ExcelTool(Tool):
    """
    Excel file manipulation tool with read, write, append, and info operations.

    Features:
    - Read Excel files (.xlsx, .xls) with range support
    - Write data to new or existing files
    - Append rows to existing files
    - Get file information and structure
    - Path security validation
    - File size limits
    - Comprehensive error handling

    Security:
    - Path traversal prevention
    - System directory protection
    - File size limits (read: 50MB, write: 10MB)
    - Format validation
    """

    max_read_size: int = Field(50 * 1024 * 1024, description="Maximum file size for reading (50MB)")
    max_write_size: int = Field(10 * 1024 * 1024, description="Maximum file size for writing (10MB)")
    allowed_extensions: List[str] = Field(default_factory=lambda: ['.xlsx', '.xls'], description="Allowed file extensions")

    def _validate_path(self, file_path: str) -> Dict[str, Any]:
        """
        Validate file path for security.

        Args:
            file_path: Path to validate

        Returns:
            Dict with 'valid' bool and optional 'error' message
        """
        # 1. Check for system sensitive directories
        forbidden_dirs = ['/etc', '/sys', '/proc', '/dev', '/boot', '/root',
                         'C:\\Windows', 'C:\\System32', '/System', '/Library']

        abs_path = os.path.abspath(file_path)

        for forbidden in forbidden_dirs:
            if abs_path.startswith(forbidden):
                return {
                    "valid": False,
                    "error": f"Access denied: Cannot access system directory {forbidden}"
                }

        # 2. Prevent path traversal
        normalized = os.path.normpath(file_path)
        if '..' in normalized.split(os.sep):
            return {
                "valid": False,
                "error": "Access denied: Path traversal detected"
            }

        # 3. Check file extension
        _, ext = os.path.splitext(file_path)
        if ext.lower() not in self.allowed_extensions:
            return {
                "valid": False,
                "error": f"Invalid file format: Only {', '.join(self.allowed_extensions)} are allowed"
            }

        return {"valid": True}

    def _check_file_size(self, file_path: str, max_size: int) -> Dict[str, Any]:
        """
        Check if file size is within limits.

        Args:
            file_path: Path to file
            max_size: Maximum allowed size in bytes

        Returns:
            Dict with 'valid' bool and optional 'error' message
        """
        if not os.path.exists(file_path):
            return {"valid": True}  # File doesn't exist yet, OK for write

        file_size = os.path.getsize(file_path)
        if file_size > max_size:
            return {
                "valid": False,
                "error": f"File size {file_size / (1024*1024):.2f}MB exceeds limit of {max_size / (1024*1024):.2f}MB"
            }

        return {"valid": True, "size": file_size}

    @retry(3, 1.0)
    def _read_excel(self, file_path: str, sheet_name: Optional[str] = None,
                   max_rows: Optional[int] = None) -> Dict[str, Any]:
        """
        Read Excel file and return data.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to read (default: first sheet)
            max_rows: Maximum rows to read (default: all)

        Returns:
            Dict with status, data, and metadata
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # Select sheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return {
                        "status": "error",
                        "error": f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}"
                    }
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            # Read data
            data = []
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                if max_rows and row_count >= max_rows:
                    break
                data.append(list(row))
                row_count += 1

            workbook.close()

            return {
                "status": "success",
                "file_path": file_path,
                "sheet_name": sheet.title,
                "rows_read": len(data),
                "columns": len(data[0]) if data else 0,
                "data": data,
                "total_rows": sheet.max_row,
                "total_columns": sheet.max_column
            }

        except Exception as e:
            LOGGER.error(f"Error reading Excel file: {str(e)}")
            return {
                "status": "error",
                "error": f"Failed to read Excel file: {str(e)}",
                "file_path": file_path
            }

    @retry(3, 1.0)
    def _write_excel(self, file_path: str, data: List[List[Any]],
                    sheet_name: str = "Sheet1", overwrite: bool = False) -> Dict[str, Any]:
        """
        Write data to Excel file.

        Args:
            file_path: Path to Excel file
            data: 2D list of data to write
            sheet_name: Sheet name
            overwrite: Whether to overwrite existing file

        Returns:
            Dict with status and operation details
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        try:
            # Check if file exists
            if os.path.exists(file_path) and not overwrite:
                return {
                    "status": "error",
                    "error": f"File already exists: {file_path}. Set overwrite=True to replace.",
                    "file_path": file_path
                }

            # Create directory if needed
            directory = os.path.dirname(file_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)

            # Create workbook and write data
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name

            for row_data in data:
                sheet.append(row_data)

            # Apply basic formatting to header row
            if data:
                for cell in sheet[1]:
                    cell.font = openpyxl.styles.Font(bold=True)

            workbook.save(file_path)
            workbook.close()

            file_size = os.path.getsize(file_path)

            return {
                "status": "success",
                "file_path": file_path,
                "sheet_name": sheet_name,
                "rows_written": len(data),
                "columns": len(data[0]) if data else 0,
                "file_size": file_size,
                "overwrite": overwrite
            }

        except Exception as e:
            LOGGER.error(f"Error writing Excel file: {str(e)}")
            return {
                "status": "error",
                "error": f"Failed to write Excel file: {str(e)}",
                "file_path": file_path
            }

    @retry(3, 1.0)
    def _append_excel(self, file_path: str, data: List[List[Any]],
                     sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Append rows to existing Excel file.

        Args:
            file_path: Path to Excel file
            data: 2D list of data to append
            sheet_name: Sheet name (default: active sheet)

        Returns:
            Dict with status and operation details
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        try:
            if not os.path.exists(file_path):
                return {
                    "status": "error",
                    "error": f"File not found: {file_path}. Use 'write' mode to create new file.",
                    "file_path": file_path
                }

            # Load existing workbook
            workbook = openpyxl.load_workbook(file_path)

            # Select sheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return {
                        "status": "error",
                        "error": f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}"
                    }
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            # Append data
            start_row = sheet.max_row + 1
            for row_data in data:
                sheet.append(row_data)

            workbook.save(file_path)
            workbook.close()

            file_size = os.path.getsize(file_path)

            return {
                "status": "success",
                "file_path": file_path,
                "sheet_name": sheet.title,
                "rows_appended": len(data),
                "start_row": start_row,
                "end_row": start_row + len(data) - 1,
                "file_size": file_size
            }

        except Exception as e:
            LOGGER.error(f"Error appending to Excel file: {str(e)}")
            return {
                "status": "error",
                "error": f"Failed to append to Excel file: {str(e)}",
                "file_path": file_path
            }

    @retry(3, 1.0)
    def _get_excel_info(self, file_path: str) -> Dict[str, Any]:
        """
        Get Excel file information.

        Args:
            file_path: Path to Excel file

        Returns:
            Dict with file metadata and structure info
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        try:
            if not os.path.exists(file_path):
                return {
                    "status": "error",
                    "error": f"File not found: {file_path}",
                    "file_path": file_path
                }

            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # Collect sheet information
            sheets_info = []
            row_counts = {}
            col_counts = {}

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                row_count = sheet.max_row
                col_count = sheet.max_column

                sheets_info.append({
                    "name": sheet_name,
                    "rows": row_count,
                    "columns": col_count
                })
                row_counts[sheet_name] = row_count
                col_counts[sheet_name] = col_count

            workbook.close()

            file_size = os.path.getsize(file_path)

            return {
                "status": "success",
                "file_path": file_path,
                "file_size": file_size,
                "file_size_mb": round(file_size / (1024 * 1024), 2),
                "total_sheets": len(workbook.sheetnames),
                "sheets": sheets_info,
                "sheet_names": workbook.sheetnames
            }

        except Exception as e:
            LOGGER.error(f"Error getting Excel info: {str(e)}")
            return {
                "status": "error",
                "error": f"Failed to get Excel info: {str(e)}",
                "file_path": file_path
            }

    def execute(self, file_path: str, mode: str, **kwargs) -> Union[str, Dict[str, Any]]:
        """
        Execute Excel operation.

        Args:
            file_path: Path to Excel file
            mode: Operation mode (read, write, append, info)
            **kwargs: Mode-specific parameters:
                - read: sheet_name, max_rows
                - write: data, sheet_name, overwrite
                - append: data, sheet_name
                - info: (no additional params)

        Returns:
            JSON string with operation result
        """
        # Validate mode
        if mode not in [m.value for m in ExcelMode]:
            return json.dumps({
                "status": "error",
                "error": f"Invalid mode: {mode}. Must be one of {[m.value for m in ExcelMode]}"
            })

        # Validate path
        path_validation = self._validate_path(file_path)
        if not path_validation["valid"]:
            return json.dumps({
                "status": "error",
                "error": path_validation["error"],
                "file_path": file_path
            })

        # Check file size for read operations
        if mode in [ExcelMode.READ.value, ExcelMode.INFO.value, ExcelMode.APPEND.value]:
            size_check = self._check_file_size(file_path, self.max_read_size)
            if not size_check["valid"]:
                return json.dumps({
                    "status": "error",
                    "error": size_check["error"],
                    "file_path": file_path
                })

        # Execute operation
        try:
            if mode == ExcelMode.READ.value:
                result = self._read_excel(
                    file_path,
                    sheet_name=kwargs.get('sheet_name'),
                    max_rows=kwargs.get('max_rows')
                )
            elif mode == ExcelMode.WRITE.value:
                data = kwargs.get('data')
                if not data:
                    return json.dumps({
                        "status": "error",
                        "error": "Parameter 'data' is required for write mode",
                        "file_path": file_path
                    })

                # Check data size
                data_size = len(json.dumps(data).encode('utf-8'))
                if data_size > self.max_write_size:
                    return json.dumps({
                        "status": "error",
                        "error": f"Data size {data_size / (1024*1024):.2f}MB exceeds write limit",
                        "file_path": file_path
                    })

                result = self._write_excel(
                    file_path,
                    data=data,
                    sheet_name=kwargs.get('sheet_name', 'Sheet1'),
                    overwrite=kwargs.get('overwrite', False)
                )
            elif mode == ExcelMode.APPEND.value:
                data = kwargs.get('data')
                if not data:
                    return json.dumps({
                        "status": "error",
                        "error": "Parameter 'data' is required for append mode",
                        "file_path": file_path
                    })
                result = self._append_excel(
                    file_path,
                    data=data,
                    sheet_name=kwargs.get('sheet_name')
                )
            elif mode == ExcelMode.INFO.value:
                result = self._get_excel_info(file_path)

            return json.dumps(result, ensure_ascii=False, indent=2)

        except Exception as e:
            LOGGER.error(f"Excel tool execution error: {str(e)}")
            return json.dumps({
                "status": "error",
                "error": f"Execution failed: {str(e)}",
                "file_path": file_path,
                "mode": mode
            })

    async def async_execute(self, file_path: str, mode: str, **kwargs) -> Union[str, Dict[str, Any]]:
        """
        Async wrapper for Excel operations.

        Args:
            file_path: Path to Excel file
            mode: Operation mode
            **kwargs: Mode-specific parameters

        Returns:
            JSON string with operation result
        """
        return await asyncio.to_thread(self.execute, file_path, mode, **kwargs)
